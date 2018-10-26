#!/usr/bin/env python3
# -*- coding: utf-8 -*-

#弥生会計取り込み用のデータを自動生成するPythonScript
#Python 3.5.1

def excel_to_yayoi():

  import openpyxl
  import os
  #直下にエクセルファイルがあるかどうかの確認し、エクセルファイルがある場合に
  for i in os.listdir():
    if i.endswith('.xlsx'):
      bookname = i

  wb = openpyxl.load_workbook(bookname)
  sheetname = wb.get_sheet_names()[1] #2番めのシート名を取得
  sheetname2 = wb.get_sheet_names()[2] #3番めのシート名を取得
  ws = wb.get_sheet_by_name(sheetname) #2番めのシートのインスタンスを取得
  ws_yayoi = wb.get_sheet_by_name(sheetname2) #3番めのシートのインスタンスを取得
  #wsからws_yayoiへデータを移行する

  rows = ws.max_row
  columns = ws.max_column #最大列

  #変数へwsのデータを挿入
  for i in range(3, ws.max_row + 1):
    num = '2000'
    datetime = ws.cell(row = i, column = 1).value #日付時刻
    date = datetime.date()#時刻を除く
    nkingaku = ws.cell(row = i, column = 2).value #入金金額
    skingaku = ws.cell(row = i, column = 3).value #出金金額
    bumon = ws.cell(row = i, column = 5).value #部門
    kamoku = ws.cell(row = i, column = 6).value #科目
    hojokamoku = ws.cell(row = i, column = 7).value #補助科目
    tekiyo = ws.cell(row = i, column = 8).value#摘要
    #弥生会計シートへ変数を入力
    ws_yayoi.cell(row = i, column = 1).value = num
    ws_yayoi.cell(row = i, column = 4).value = date

    if nkingaku is not None and skingaku is None: #預金借方
        num = [5,6,7,8,9,11,12,13,14,15,17,20,25]
        text = ['普通預金',
                'みずほ銀行虎ノ門',
                '全社',
                '対象外',
                nkingaku,
                kamoku,
                hojokamoku,
                bumon,
                '込',
                nkingaku,
                tekiyo,
                0,
                'no']
        for n, v in zip(num, text):
            ws_yayoi.cell(row=i, column=n).value = v

    elif nkingaku is None and skingaku is not None:#預金貸方
        num = [5,6,7,8,9,11,12,13,14,15,17,20,25]
        text = [kamoku,
                hojokamoku,
                bumon,
                '込',
                skingaku,
                '普通預金',
                'みずほ銀行虎ノ門',
                '全社',
                '対象外',
                skingaku,
                tekiyo,
                0,
                'no']
        for n, v in zip(num, text):
            ws_yayoi.cell(row=i, column=n).value = v

  wb.save(filename = bookname)

def yayoi_to_csv():

  import openpyxl
  import os
  #直下にエクセルファイルがあるかどうかの確認し、エクセルファイルがある場合に
  #ファイル名を変数に代入
  for i in os.listdir():
    if i.endswith('.xlsx'):
      bookname = i

  #エクセルファイルを読み込み、弥生会計の必要な項目の変数に値を代入
  wb = openpyxl.load_workbook(bookname)
  sheetname = wb.get_sheet_names()[1] #2番めのシート名を取得
  sheetname2 = wb.get_sheet_names()[2] #3番めのシート名を取得
  ws = wb.get_sheet_by_name(sheetname) #2番めのシートのインスタンスを取得
  ws_yayoi = wb.get_sheet_by_name(sheetname2) #3番めのシートのインスタンスを取得
  #wsからws_yayoiへデータを移行する

  rows = ws_yayoi.max_row #最大行
  columns = ws_yayoi.max_column #最大列

  #弥生会計形式のシートをcsv形式に変換
  import csv
  import codecs

  #弥生会計形式のシートをリストに変換
  body = [[ws_yayoi.cell(row = i, column = v).value for v in range(1, 26)]for i in range(3, ws.max_row + 1)]

  #時刻除外
  for i in body:
    i[3] = i[3].strftime('%Y%m%d')

  #ファイルを書き込みモードでオープン&書き込みの上、クローズ
  with codecs.open('ToYayoi.csv', 'w', 'cp932') as f:
    writer = csv.writer(f)#writerオブジェクトを作成
    writer.writerows(body)#二次元配列を代入


if __name__ == '__main__':
  excel_to_yayoi()
  yayoi_to_csv()
